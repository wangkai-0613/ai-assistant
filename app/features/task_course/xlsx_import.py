"""解析教务系统导出的按周分块 XLSX 课表。"""

from __future__ import annotations

import re
import uuid
from datetime import date, datetime, timedelta

from openpyxl import load_workbook

from app.core.contracts import Course

_WEEK_PATTERN = re.compile(r"^第\d+周$")
_DATE_RANGE_PATTERN = re.compile(r"(\d{4}-\d{2}-\d{2})\s*至\s*(\d{4}-\d{2}-\d{2})")
SECTION_TIMES = {
    "1-2": ("08:00", "09:40"),
    "3-4": ("10:00", "11:40"),
    "5-6": ("14:00", "15:40"),
    "7-8": ("16:00", "17:40"),
    "9-12": ("19:00", "22:00"),
    "1-4": ("08:00", "11:40"),
    "5-8": ("14:00", "17:40"),
}


def parse_xlsx(path: str) -> list[Course]:
    """读取首个工作表并返回带具体日期的课程。"""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = [tuple(row) for row in workbook.active.iter_rows(values_only=True)]
    finally:
        workbook.close()
    return _parse_week_blocks(rows)


def _parse_week_blocks(rows: list[tuple]) -> list[Course]:
    courses: list[Course] = []
    starts = [
        index
        for index, row in enumerate(rows)
        if row and _WEEK_PATTERN.match(_text(row[0]))
    ]
    for block_index, row_index in enumerate(starts):
        block_end = starts[block_index + 1] if block_index + 1 < len(starts) else len(rows)
        week_start = _week_start(rows, row_index)
        if week_start is None:
            continue
        width = min(max((len(row) for row in rows[row_index:block_end]), default=0), 8)
        for column in range(1, width):
            courses.extend(_parse_day(rows, row_index, block_end, column, week_start))
    return courses


def _week_start(rows: list[tuple], row_index: int) -> date | None:
    if row_index + 1 >= len(rows) or not rows[row_index + 1]:
        return None
    match = _DATE_RANGE_PATTERN.search(_text(rows[row_index + 1][0]))
    return datetime.strptime(match.group(1), "%Y-%m-%d").date() if match else None


def _parse_day(rows, start: int, end: int, column: int, week_start: date) -> list[Course]:
    courses: list[Course] = []
    cursor = start
    while cursor + 2 < end:
        section = _prefixed_value(rows, cursor, column, "节次：")
        name = _prefixed_value(rows, cursor + 1, column, "课程名称：")
        times = SECTION_TIMES.get(section)
        if name and times is not None:
            room = _prefixed_value(rows, cursor + 2, column, "上课教室：", allow_empty=True)
            course_date = week_start + timedelta(days=column - 1)
            courses.append(
                Course(
                    id=uuid.uuid4().hex,
                    name=name,
                    weekday=course_date.isoweekday(),
                    start_time=times[0],
                    end_time=times[1],
                    room=room,
                    course_date=course_date,
                )
            )
            cursor += 3
        else:
            cursor += 1
    return courses


def _prefixed_value(rows, row: int, column: int, prefix: str, allow_empty=False) -> str:
    if row >= len(rows) or column >= len(rows[row]):
        return ""
    value = _text(rows[row][column])
    if not value.startswith(prefix):
        return ""
    result = value[len(prefix) :].strip()
    return result if result or allow_empty else ""


def _text(value) -> str:
    return "" if value is None else str(value).strip()
